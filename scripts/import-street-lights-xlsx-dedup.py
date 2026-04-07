#!/usr/bin/env python3
from __future__ import annotations

import json
import math
import os
from pathlib import Path
from typing import Iterable

import openpyxl
import requests


def load_local_env() -> None:
    env_path = Path(__file__).resolve().parents[1] / ".env.local"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s or s.startswith("#") or "=" not in s:
            continue
        k, v = s.split("=", 1)
        k = k.strip()
        v = v.strip().strip("'").strip('"')
        if k and k not in os.environ:
            os.environ[k] = v


def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371000.0
    p = math.pi / 180.0
    a = (
        0.5
        - math.cos((lat2 - lat1) * p) / 2
        + math.cos(lat1 * p) * math.cos(lat2 * p) * (1 - math.cos((lon2 - lon1) * p)) / 2
    )
    return r * 2 * math.asin(math.sqrt(a))


def read_xlsx_points(path: Path) -> list[tuple[float, float]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    points: list[tuple[float, float]] = []

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        nums = [v for v in row if isinstance(v, (int, float))]
        lat = None
        lng = None
        for n in nums:
            f = float(n)
            # Korea lat/lng heuristic
            if 33.0 <= f <= 39.5 and lat is None:
                lat = f
            elif 124.0 <= f <= 132.5 and lng is None:
                lng = f
        if lat is None or lng is None:
            continue
        points.append((lat, lng))
    return points


def batch(items: list[dict], n: int) -> Iterable[list[dict]]:
    for i in range(0, len(items), n):
        yield items[i : i + n]


def main() -> int:
    load_local_env()

    supabase_url = (
        os.environ.get("SUPABASE_URL")
        or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
        or ""
    ).rstrip("/")
    key = (os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or "").strip()
    xlsx_path = Path(
        os.environ.get(
            "STREET_LIGHTS_XLSX",
            r"C:\Users\PC\Desktop\종프\가로등공공데이터.xlsx",
        )
    )
    threshold_m = float(os.environ.get("STREET_LIGHTS_DEDUP_M", "25"))

    if not supabase_url or not key:
        print("Missing NEXT_PUBLIC_SUPABASE_URL/SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")
        return 1
    if not xlsx_path.exists():
        print(f"XLSX not found: {xlsx_path}")
        return 1

    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    # 1) Existing DB points
    sel_url = f"{supabase_url}/rest/v1/street_lights?select=id,light_no,lat,lng&limit=20000"
    resp = requests.get(sel_url, headers=headers, timeout=60)
    resp.raise_for_status()
    existing_rows = resp.json()
    existing_points: list[tuple[float, float]] = [
        (float(r["lat"]), float(r["lng"]))
        for r in existing_rows
        if r.get("lat") is not None and r.get("lng") is not None
    ]
    max_light_no = 0
    max_id = 0
    for r in existing_rows:
        rid = r.get("id")
        if isinstance(rid, int) and rid > max_id:
            max_id = rid
        v = r.get("light_no")
        if isinstance(v, int) and v > max_light_no:
            max_light_no = v

    # 2) XLSX points
    input_points = read_xlsx_points(xlsx_path)
    if not input_points:
        print("No valid lat/lng rows in XLSX.")
        return 1

    # 3) Dedup by 25m against existing db points (+ accepted new points)
    accepted_points: list[tuple[float, float]] = []
    skipped = 0

    compare_pool = existing_points.copy()
    for lat, lng in input_points:
        too_close = False
        for elat, elng in compare_pool:
            if haversine_m(lat, lng, elat, elng) <= threshold_m:
                too_close = True
                break
        if too_close:
            skipped += 1
            continue
        accepted_points.append((lat, lng))
        compare_pool.append((lat, lng))

    if not accepted_points:
        print(
            f"Done. input={len(input_points)}, existing={len(existing_points)}, "
            f"inserted=0, skipped_within_{int(threshold_m)}m={skipped}"
        )
        return 0

    payload_rows = []
    for i, (lat, lng) in enumerate(accepted_points, start=1):
        payload_rows.append(
            {
                "id": max_id + i,
                "light_no": max_light_no + i,
                "lat": lat,
                "lng": lng,
            }
        )

    ins_url = f"{supabase_url}/rest/v1/street_lights"
    inserted = 0
    for chunk in batch(payload_rows, 500):
        ir = requests.post(ins_url, headers=headers, data=json.dumps(chunk), timeout=60)
        if ir.status_code >= 300:
            print(f"Insert failed: {ir.status_code} {ir.text}")
            return 1
        inserted += len(chunk)

    print(
        f"Done. input={len(input_points)}, existing={len(existing_points)}, "
        f"inserted={inserted}, skipped_within_{int(threshold_m)}m={skipped}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
